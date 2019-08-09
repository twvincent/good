#!/usr/bin/python3

import openpyxl
data1, data2 = [], []
filename = 'pyxls.xlsx'

def printsheet(worksheet) :
    '''display sheet data'''
    for row in worksheet.rows :
        for cell in row :
            print(cell.value, end='\t')
        print()
    return

def resetdata() :
    ''' remove data from  data1[] & data2[] '''
    global data1, data2
    data1 = []
    data2 = []
    return

def append2end(c2,row2) :
    global data1, data2
    x = c2.row
    for y in range(3,6) :
        if (sheet2.cell(row=x, column=y).value == None) :
            sheet2.cell(row=x, column=y).value = data1[1]
            return
    return


# open excel file
book = openpyxl.load_workbook(filename, data_only=True)
sheet = book.active

# delete sheet('new') if already exist
sheet2 = book.create_sheet('new')
if len(sheet2.rows) > 0 :
    book.remove_sheet(sheet2)
    sheet2 = book.create_sheet('new')

# copy data to sheet2
for row1 in sheet.rows :
    for c1 in row1 :
        data1.append(c1.value) # store row data to list

    # fetch sheet2 data
    exist = False
    for row2 in sheet2.rows :
        for c2 in row2 :
            data2.append(c2.value)

        # already exist in sheet2?
        if (data1[0] == data2[0]) :
            append2end(c2,row2) 
            exist = True
            resetdata()
            break
        data2 = []
        
    if (not exist) :
        sheet2.append(data1)
        resetdata()

#printsheet(sheet2)
book.save(filename)
print('Done')
